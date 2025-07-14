# ------------------- CIF Charges Entry UI -------------------
# LCL Destination Charges Comparison Calculator with Save & History Tabs
# ----------------------------------------------------------------------
import streamlit as st
import pandas as pd
import numpy as np
import os
import re
from io import BytesIO
import zipfile
import openpyxl

# ----------------------------------------------------------------------
# 0.  Page setup (MUST be first Streamlit call)
# ----------------------------------------------------------------------
st.set_page_config(
    page_title="LCL Destination Charges Comparison Calculator",
    page_icon="📦",
    layout="wide",
)

# Ensure data directories exist ------------------------------------------------
DATA_DIR = "Data"
SAVED_DIR = os.path.join(DATA_DIR, "Saved")
EXCHANGE_PATH = os.path.join(DATA_DIR, "Exchange Rates.xlsx")
os.makedirs(SAVED_DIR, exist_ok=True)

# ---------------------------------------------------------------------------------------
# 1.  Helper – Load & cache exchange‑rate table
# ---------------------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def load_exchange_rates() -> pd.DataFrame:
    """Returns the exchange‑rate DataFrame directly from Excel."""
    return pd.read_excel(EXCHANGE_PATH)

def save_exchange_rates(df: pd.DataFrame):
    df.to_excel(EXCHANGE_PATH, index=False)

exchange_df = load_exchange_rates()

@st.cache_data(show_spinner=False)
def get_currency_list(df: pd.DataFrame):
    return sorted(df["Currency"].dropna().unique().tolist())

currency_options = get_currency_list(exchange_df)

# ------------------------------------------------------------------
# 6.  Comparison engine
# ------------------------------------------------------------------
def nom(con_cbm,con_bl,freight_cost,market_rate,nomination_rate,nomination_cbm,nomination_bl,rebate_cbm,rebate_bl,rebate_per_container,tran_cbm_f,tran_pro_per_cbm_f):
    free_hand_volume = float(con_cbm) - float(nomination_cbm)
    free_hand_bl = int(con_bl-nomination_bl)

    pro_free_hand = (free_hand_volume*market_rate)+(free_hand_volume*rebate_cbm)-(free_hand_volume*freight_cost)+(free_hand_bl*rebate_bl)
    pro_nomination = (nomination_rate-freight_cost)*nomination_cbm

    pro_sum = pro_free_hand+pro_nomination+rebate_per_container+(tran_cbm_f*tran_pro_per_cbm_f)

    return free_hand_volume,free_hand_bl,pro_free_hand,pro_nomination,pro_sum

def agent_compare(df,nom_df,input_dict,exchange_df):
    money_cols = ['Per CBM', 'Per Ton', 'Minimum', 'Maximum', 'Per BL']

    loadability_20_f = input_dict["20'STD"][0]
    box_rate_20_f = input_dict["20'STD"][1]
    num_bl_20_f = input_dict["20'STD"][2]
    market_rate_20_f = input_dict["20'STD"][3]
    tran_cbm_20_f = input_dict["20'STD"][4]
    tran_num_bl_20_f = input_dict["20'STD"][5]
    tran_pro_per_cbm_20_f = input_dict["20'STD"][6]
    con_cbm_20 = float(loadability_20_f)-float(tran_cbm_20_f)
    freight_cost_20 = float(box_rate_20_f)/float(loadability_20_f)
    con_bl_20 = float(num_bl_20_f)-float(tran_num_bl_20_f)

    loadability_40_f = input_dict["40'STD"][0]
    box_rate_40_f = input_dict["40'STD"][1]
    num_bl_40_f = input_dict["40'STD"][2]
    market_rate_40_f = input_dict["40'STD"][3]
    tran_cbm_40_f = input_dict["40'STD"][4]
    tran_num_bl_40_f = input_dict["40'STD"][5]
    tran_pro_per_cbm_40_f = input_dict["20'STD"][6]
    con_cbm_40 = float(loadability_40_f)-float(tran_cbm_40_f)
    freight_cost_40 = float(box_rate_40_f)/float(loadability_40_f)
    con_bl_40 = float(num_bl_40_f)-float(tran_num_bl_40_f)

    # Clean numeric columns
    df[money_cols] = (df[money_cols]
                    .replace(r'^\s*$', np.nan, regex=True)
                    .apply(pd.to_numeric, errors='coerce')
                    .fillna(0))

    # Currency → USD map
    rate_map = dict(zip(exchange_df['Currency'],
                        exchange_df['Exchange Rate to USD'].astype(float)))
    rate_map.setdefault('USD', 1.0)

    # Output rows
    rows_out = []
    nomination_out = []
    for agent, grp in df.groupby('Agent Name', sort=False):
        rebate_df  = grp[grp['Description'] == 'Rebate']
        remarks_df = grp[grp['Description'] == 'Remarks']
        charge_df  = grp[~grp['Description'].isin(['Rebate', 'Remarks'])]

        remark = remarks_df['Currency'].iloc[0] if not remarks_df.empty else ""
        nomination_rate = nom_df[nom_df['Agent Name'] == agent]["Nomination Rate"].values[0]
        nomination_cbm = nom_df[nom_df['Agent Name'] == agent]["Nomination CBM"].values[0]
        nomination_bl = nom_df[nom_df['Agent Name'] == agent]["Nomination BL"].values[0]

        # Rebates
        if rebate_df.empty:
            rebate_cbm = rebate_per_ton = rebate_bl = rebate_per_container = 0.0
        else:
            r_cur = rebate_df.iloc[0]['Currency']
            r_rate = rate_map.get(r_cur, np.nan)
            rebate_cbm     = rebate_df.iloc[0]['Per CBM'] * r_rate if not np.isnan(r_rate) else 0
            rebate_bl      = rebate_df.iloc[0]['Per BL']  * r_rate if not np.isnan(r_rate) else 0
            rebate_per_ton = rebate_df.iloc[0]['Per Ton'] * r_rate if not np.isnan(r_rate) else 0
            rebate_per_container = float(rebate_df.iloc[0]["Per Container"]) * r_rate if not np.isnan(r_rate) else 0

        free_hand_volume_20,free_hand_bl_20,pro_free_hand_20,pro_nomination_20,pro_sum_20 = nom(con_cbm_20,con_bl_20,freight_cost_20,market_rate_20_f,
                                                                                                nomination_rate,nomination_cbm,nomination_bl,rebate_cbm,rebate_bl,
                                                                                                rebate_per_container,tran_cbm_20_f,tran_pro_per_cbm_20_f)
        free_hand_volume_40,free_hand_bl_40,pro_free_hand_40,pro_nomination_40,pro_sum_40 = nom(con_cbm_40,con_bl_40,freight_cost_40,market_rate_40_f,
                                                                                                nomination_rate,nomination_cbm,nomination_bl,rebate_cbm,rebate_bl,
                                                                                                rebate_per_container,tran_cbm_40_f,tran_pro_per_cbm_40_f)
        
        now_row1 = {"Agent Name":agent,"Container Type":"20'STD","Box Rate":box_rate_20_f,"Total Loadability":loadability_20_f,
                    "Freight Cost":freight_cost_20,"Total Number of BLs":num_bl_20_f,"Market Rate":market_rate_20_f,
                    "Nomination Rate":nomination_rate,"Transhipment CBM":tran_cbm_20_f,"Transhipment Number of BLs":tran_num_bl_20_f,
                    "Transhipment Profitability Per CBM":tran_pro_per_cbm_20_f,"Rebate Per CBM":rebate_cbm,"Rebate Per BL":rebate_bl,
                    "Rebate Per Container":rebate_per_container,"Nomination CBM":nomination_cbm,
                    "Nomination BL":nomination_bl,"Considered CBM":con_cbm_20,"Considered BLs":con_bl_20,
                    "Free Hand CBM":free_hand_volume_20,"Free Hand BL":free_hand_bl_20,"Profitability on Free Hand":pro_free_hand_20,
                    "Profitability on Nomination":pro_nomination_20,"Sum of Profitability":pro_sum_20}
        
        now_row2 = {"Agent Name":agent,"Container Type":"40'STD","Box Rate":box_rate_40_f,"Total Loadability":loadability_40_f,
                    "Freight Cost":freight_cost_40,"Total Number of BLs":num_bl_40_f,"Market Rate":market_rate_40_f,
                    "Nomination Rate":nomination_rate,"Transhipment CBM":tran_cbm_40_f,"Transhipment Number of BLs":tran_num_bl_40_f,
                    "Transhipment Profitability Per CBM":tran_pro_per_cbm_20_f,"Rebate Per CBM":rebate_cbm,"Rebate Per BL":rebate_bl,
                    "Rebate Per Container":rebate_per_container,"Nomination CBM":nomination_cbm,
                    "Nomination BL":nomination_bl,"Considered CBM":con_cbm_40,"Considered BLs":con_bl_40,
                    "Free Hand CBM":free_hand_volume_40,"Free Hand BL":free_hand_bl_40,"Profitability on Free Hand":pro_free_hand_40,
                    "Profitability on Nomination":pro_nomination_40,"Sum of Profitability":pro_sum_40}
        
        nomination_out.extend([now_row1, now_row2])

        # Total charges
        totals = charge_df.apply(
            lambda row: row[money_cols] * rate_map.get(row['Currency'], np.nan),
            axis=1
        ).sum()
        tot_cbm, tot_bl, tot_ton = totals['Per CBM'], totals['Per BL'], totals['Per Ton']

        row1 = {"Agent Name": agent, "Remarks": remark, "Type": "Destination Charges"}
        row2 = {"Agent Name": agent, "Remarks": remark, "Type": "Fixed Charges (BL)"}
        row3 = {"Agent Name": agent, "Remarks": remark, "Type": "Rebate (CBM or Ton)"}
        row4 = {"Agent Name": agent, "Remarks": remark, "Type": "Rebate (BL)"}
        row5 = {"Agent Name": agent, "Remarks": remark, "Type": "Net Charges"}

        for n in range(1, 31):
            tpc = tot_cbm * n
            tpt = tot_ton * (n / 2)  # as per your logic: ton weight = CBM / 2

            if tpc > tpt:
                con = tpc
                rcon = rebate_cbm * n
            else:
                con = tpt
                rcon = rebate_per_ton * (n / 2)

            dest_chg = tot_bl + con
            net = dest_chg - rcon - rebate_bl

            row1[f"CBM {n}"] = round(con, 2)
            row2[f"CBM {n}"] = round(tot_bl, 2)
            row3[f"CBM {n}"] = round(rcon, 2)
            row4[f"CBM {n}"] = round(rebate_bl, 2)
            row5[f"CBM {n}"] = round(net, 2)

        rows_out.extend([row1, row2, row3, row4, row5])

    comp_df = pd.DataFrame(rows_out)
    nomination_df = pd.DataFrame(nomination_out)
    return comp_df,nomination_df

# ==============================================================================
# MAIN NAVIGATION TABS
# ==============================================================================
main_tabs = st.tabs([
    "📊 Comparison Calculator",
    "📂 Saved Comparisons",
    "💱 Exchange Rates",
    "🚢 Port of Discharge"
])


# ==============================================================================
# TAB 1: COMPARISON CALCULATOR
# ==============================================================================
with main_tabs[0]:
    st.title("LCL Destination Charges Comparison Calculator")

    locations_df = pd.read_excel(r"Data/locations.xlsx", sheet_name="POD locations")
    pod_list = sorted(locations_df['POD'].dropna().unique())
    # Create two columns for POL and POD dropdowns
    col1, col2 = st.columns(2)

    with col1:
        pol = st.selectbox("**Port of Loading (POL)**", ["Nhava Sheva"],key="pol")

    with col2:
        pod = st.selectbox("**Port of Discharge (POD)**", pod_list,key="pod")
    # ------------------------------------------------------------------
    # 2.  Container‑level inputs
    # ------------------------------------------------------------------
    with st.expander("***📦 20' STD Information***", expanded=True):
        c1, c2, c3, c4, c5 = st.columns(5)
        box_rate_20    = c1.text_input("**Box Rate (USD)**", "0", key="box_rate_20")
        loadability_20 = c2.text_input("**Loadability (numeric)**", "0", key="load_20")
        num_bl_20      = c3.text_input("**Number of BLs (numeric)**", "0", key="num_bl_20")
        market_rate_20 = c4.text_input("**Market Rate (USD)**", "0", key="mkt_rate_20")
        
        # Freight cost display (in new column next to box rate & loadability)
        try:
            box_20 = float(box_rate_20)
            load_20 = float(loadability_20)
            if load_20 > 0:
                freight_cost_20 = box_20 / load_20
                c5.metric("📉 Freight Cost Per CBM", f"${freight_cost_20:.2f}")
            else:
                c5.write("Enter loadability > 0")
        except ValueError:
            c5.write("Waiting for valid numbers")

        # Transhipment
        st.markdown("**Transhipment**")
        t1, t2, t3, t4, t5 = st.columns(5)
        tran_cbm_20    = t1.text_input("**CBM (numeric)**", "0", key="tran_cbm_20")
        tran_num_bl_20 = t2.text_input("**# of BLs (numeric)**", "0", key="tran_num_bl_20")
        tran_pro_per_cbm_20 = t3.text_input("**Profitability Per CBM**", "0", key="tran_pro_per_cbm_20")


    with st.expander("***📦 40' STD Information***", expanded=True):
        c1, c2, c3, c4, c5 = st.columns(5)
        box_rate_40       = c1.text_input("**Box Rate (USD)**", "0", key="box_rate_40")
        loadability_40    = c2.text_input("**Loadability (numeric)**", "0", key="load_40")
        num_bl_40         = c3.text_input("**Number of BLs (numeric)**", "0", key="num_bl_40")
        market_rate_40    = c4.text_input("**Market Rate (USD)**", "0", key="mkt_rate_40")
        # Freight cost display (in new column next to box rate & loadability)
        try:
            box_40 = float(box_rate_40)
            load_40 = float(loadability_40)
            if load_40 > 0:
                freight_cost_40 = box_40 / load_40
                c5.metric("📉 Freight Cost Per CBM", f"${freight_cost_40:.2f}")
            else:
                c5.write("Enter loadability > 0")
        except ValueError:
            c5.write("Waiting for valid numbers")

        st.markdown("**Transhipment**")
        t1,t2,t3,t4,t5 = st.columns(5)
        tran_cbm_40    = t1.text_input("**CBM (numeric)**", "0", key="tran_cbm_40")
        tran_num_bl_40 = t2.text_input("**# of BLs (numeric)**", "0", key="tran_num_bl_40")
        tran_pro_per_cbm_40 = t3.text_input("**Profitability Per CBM**", "0", key="tran_pro_per_cbm_40")

    # ------------------------------------------------------------------
    # 3.  Session‑state setup for dynamic agents
    # ------------------------------------------------------------------
    if "agent_ids" not in st.session_state:
        st.session_state.agent_ids  = [1]
        st.session_state.agent_names = {1: "Agent 1"}

    if st.button("➕ Add Agent"):
        new_id = max(st.session_state.agent_ids) + 1
        st.session_state.agent_ids.append(new_id)
        st.session_state.agent_names[new_id] = f"Agent {new_id}"

    def delete_agent(agent_id):
        st.session_state.agent_ids.remove(agent_id)
        st.session_state.agent_names.pop(agent_id, None)

    # ------------------------------------------------------------------
    # 4.  Data‑collector helpers
    # ------------------------------------------------------------------
    def extract_agent_data() -> pd.DataFrame:
        rows = []
        nom_rows = []

        for agent_id in st.session_state.agent_ids:
            agent_name = st.session_state.get(f"agent_name_{agent_id}", f"Agent {agent_id}")
            nomination_rate = st.session_state.get(f"nom_support_rate_{agent_id}", 0)
            nomination_cbm = st.session_state.get(f"nom_support_cbm_{agent_id}", 0)
            nomination_bl = st.session_state.get(f"nom_support_bl_{agent_id}", 0)

            nom_rows.append({
                "Agent Name": agent_name,
                "Nomination Rate": nomination_rate,
                "Nomination CBM": nomination_cbm,
                "Nomination BL": nomination_bl
            })

            # 🔁 Extract dynamic number of charge head rows
            num_rows = st.session_state.get(f"{agent_id}_num_charge_rows", 8)
            for i in range(1, num_rows + 1):
                desc = st.session_state.get(f"{agent_id}_desc_{i}", "")
                if desc.strip() == "":
                    continue  # skip empty rows

                rows.append({
                    "Agent Name": agent_name,
                    "Description": desc,
                    "Currency":    st.session_state.get(f"{agent_id}_currency_{i}", ""),
                    "Per CBM":     st.session_state.get(f"{agent_id}_cbm_{i}", ""),
                    "Per Ton":     st.session_state.get(f"{agent_id}_ton_{i}", ""),
                    "Minimum":     st.session_state.get(f"{agent_id}_min_{i}", ""),
                    "Maximum":     st.session_state.get(f"{agent_id}_max_{i}", ""),
                    "Per BL":      st.session_state.get(f"{agent_id}_bl_{i}", ""),
                    "Vat(%)" :     st.session_state.get(f"{agent_id}_vat_{i}",""),
                    "Per Container": ""
                })

            # 📝 Notes row (Charge Head 9 Notes)
            rows.append({
                "Agent Name": agent_name,
                "Description": "Remarks",
                "Currency":    st.session_state.get(f"{agent_id}_desc_9_notes", ""),
                "Per CBM": "", "Per Ton": "", "Minimum": "", "Maximum": "", "Per BL": "", "Vat(%)" : "", "Per Container": ""
            })

            # 🎯 Rebate row
            rows.append({
                "Agent Name": agent_name,
                "Description": "Rebate",
                "Currency": st.session_state.get(f"{agent_id}_rebate_currency", ""),
                "Per CBM":   st.session_state.get(f"{agent_id}_rebate_cbm", ""),
                "Per Ton":   st.session_state.get(f"{agent_id}_rebate_ton", ""),
                "Minimum": "", "Maximum": "",
                "Per BL":    st.session_state.get(f"{agent_id}_rebate_bl", ""), "Vat(%)" : "",
                "Per Container": st.session_state.get(f"{agent_id}_rebate_container", "")
            })

        df = pd.DataFrame(rows)
        agent_df = df[df["Description"].fillna("").str.strip() != ""]  # remove blank desc rows
        nomination_df = pd.DataFrame(nom_rows)

        return agent_df, nomination_df


    # ------------------------------------------------------------------
    # 5.  Agent‑entry form UI
    # ------------------------------------------------------------------
    def agent_form(agent_id: int):
        c1, c2 = st.columns([5, 1])
        with c1:
            st.text_input("***Agent Name***", key=f"agent_name_{agent_id}",
                          value=st.session_state.agent_names[agent_id])
        with c2:
            if st.button("❌", key=f"del_{agent_id}"):
                delete_agent(agent_id)
                st.rerun()

        n1,n2,n3,n4,n5 = st.columns(5)
        n1.number_input(
            "**Nomination Rate(USD)**",
            min_value=0.0,
            step=0.1,
            key=f"nom_support_rate_{agent_id}",
            value=float(st.session_state.get(f"nom_support_rate_{agent_id}", 0.0))
        )
        n2.number_input(
            "**Nomination CBM**",
            min_value=0.0,
            step=0.1,
            key=f"nom_support_cbm_{agent_id}",
            value=float(st.session_state.get(f"nom_support_cbm_{agent_id}", 0.0))
        )

        n3.number_input(
            "**Nomination BL**",
            min_value=0,
            step=1,
            key=f"nom_support_bl_{agent_id}",
            value=int(st.session_state.get(f"nom_support_bl_{agent_id}", 0))
        )

        # Destination Charges Header (Fixed)
        st.markdown("***Destination Charges (CIF)***")
        head_cols = st.columns([3, 1, 1, 1, 1, 1, 1, 1])  # Fixed to 8 columns
        for col, h in zip(head_cols,
                            ["Charge Head", "Currency", "Per CBM", "Per Ton",
                            "Minimum", "Maximum", "Per BL", "Vat(%)"]):
            col.markdown(f"**{h}**")

        if f"{agent_id}_num_charge_rows" not in st.session_state:
            st.session_state[f"{agent_id}_num_charge_rows"] = 8

        for i in range(1, st.session_state[f"{agent_id}_num_charge_rows"] + 1):
            cols = st.columns([3, 1, 1, 1, 1, 1, 1, 1])
            cols[0].text_input("", key=f"{agent_id}_desc_{i}",
                            label_visibility="collapsed", placeholder=f"Charge Head {i}")
            cols[1].selectbox("", currency_options, key=f"{agent_id}_currency_{i}",
                label_visibility="collapsed",
                index=currency_options.index("USD")
                if "USD" in currency_options else 0)
            cols[2].text_input("", key=f"{agent_id}_cbm_{i}", label_visibility="collapsed")
            cols[3].text_input("", key=f"{agent_id}_ton_{i}", label_visibility="collapsed")
            cols[4].text_input("", key=f"{agent_id}_min_{i}", label_visibility="collapsed")
            cols[5].text_input("", key=f"{agent_id}_max_{i}", label_visibility="collapsed")
            cols[6].text_input("", key=f"{agent_id}_bl_{i}", label_visibility="collapsed")
            cols[7].text_input("", key=f"{agent_id}_vat_{i}", label_visibility="collapsed")

        # Add Charge Head button
        if st.button("➕ Add Charge Head", key=f"add_charge_head_{agent_id}"):
            st.session_state[f"{agent_id}_num_charge_rows"] += 1

        st.text_input("Charge Head Notes", key=f"{agent_id}_desc_notes",
                      placeholder="If Cartons, …")

        st.markdown("***Rebates***")

        rebate_cols = st.columns(5)
        rebate_headers = ["Currency", "Per CBM", "Per Ton", "Per BL","Per Container"]
        for col, header in zip(rebate_cols, rebate_headers):
            col.markdown(f"**{header}**")

        r1, r2, r3, r4, r5 = st.columns(5)
        r1.selectbox("", currency_options,
                     key=f"{agent_id}_rebate_currency",
                     label_visibility="collapsed",
                     index=currency_options.index("USD") if "USD" in currency_options else 0)
        r2.text_input("", key=f"{agent_id}_rebate_cbm", label_visibility="collapsed")
        r3.text_input("", key=f"{agent_id}_rebate_ton", label_visibility="collapsed")
        r4.text_input("", key=f"{agent_id}_rebate_bl", label_visibility="collapsed")
        r5.text_input("", key=f"{agent_id}_rebate_container", label_visibility="collapsed")

    # render each agent tab
    tabs = st.tabs([f"Agent {aid}" for aid in st.session_state.agent_ids])
    for tab, aid in zip(tabs, st.session_state.agent_ids):
        with tab:
            agent_form(aid)

    st.markdown("""
    <div style="color: red; font-weight: bold;">
    ⚠️ Please note: <u>VAT(%) is not included</u> in any of the comparison or calculation logic.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 🛠️ Actions")
    calc_btn, dl_placeholder, save_placeholder = st.columns([1, 1, 1])

    # 7‑A Calculate
    if calc_btn.button("🧮 Calculate"):
        try:
            box_rate_20_f = float(box_rate_20)
            loadability_20_f = float(loadability_20)
            num_bl_20_f = float(num_bl_20)
            market_rate_20_f = float(market_rate_20)
            tran_cbm_20_f = float(tran_cbm_20)
            tran_num_bl_20_f = float(tran_num_bl_20)
            tran_pro_per_cbm_20_f = float(tran_pro_per_cbm_20)

            box_rate_40_f = float(box_rate_40)
            loadability_40_f = float(loadability_40)
            num_bl_40_f = float(num_bl_40)
            market_rate_40_f = float(market_rate_40)
            tran_cbm_40_f = float(tran_cbm_40)
            tran_num_bl_40_f = float(tran_num_bl_40)
            tran_pro_per_cbm_40_f = float(tran_pro_per_cbm_40)

        except ValueError:
            st.error("Loadability, Box Rate, and Origin Charges must be numeric.")
            st.stop()

        input_dict = {
                "20'STD": [
                    loadability_20_f, box_rate_20_f, num_bl_20_f, market_rate_20_f,
                    tran_cbm_20_f, tran_num_bl_20_f, tran_pro_per_cbm_20_f
                ],
                "40'STD": [
                    loadability_40_f, box_rate_40_f, num_bl_40_f, market_rate_40_f,
                    tran_cbm_40_f, tran_num_bl_40_f, tran_pro_per_cbm_40_f
                ]
        }
        in_df, nom_df  = extract_agent_data()
        comp_df,nomination_df = agent_compare(in_df,nom_df,input_dict,exchange_df)

        st.session_state["container_info"] = pd.DataFrame({
                "Field": [
                    "POL","POD","Loadability", "Box Rate (USD)", "Number of BLs", "Market Rate (USD)",
                    "Transhipment CBM", "Transhipment Number of BLs", "Transhipment Profitability Per CBM"
                ],
                "20'STD": [
                    pol,pod,loadability_20_f, box_rate_20_f, num_bl_20_f, market_rate_20_f,
                    tran_cbm_20_f, tran_num_bl_20_f, tran_pro_per_cbm_20_f
                ],
                "40'STD": [
                    pol,pod,loadability_40_f, box_rate_40_f, num_bl_40_f, market_rate_40_f,
                    tran_cbm_40_f, tran_num_bl_40_f, tran_pro_per_cbm_40_f
                ]
            })

        st.session_state["last_input_df"]  = in_df
        st.session_state["last_nom_df"] = nom_df
        st.session_state["last_result_df"] = comp_df
        st.session_state["last_nomination_df"] = nomination_df


        st.success("Calculation complete.")
        st.dataframe(comp_df)
        st.dataframe(nom_df)
        st.dataframe(nomination_df)

    # 7‑B Download (only if data exists)
    def to_safe_sheet(name: str) -> str:
        # Trim to 31 chars, remove forbidden chars
        name = re.sub(r"[\[\]\*:/\\?]", "", name)[:31]
        return name or "Sheet"

    if all(k in st.session_state for k in ("container_info", "last_input_df", "last_result_df","last_nomination_df")):
        with dl_placeholder:
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
                st.session_state["container_info"].to_excel(writer, sheet_name="Info", index=False)

                for agent, grp in st.session_state["last_input_df"].groupby("Agent Name", sort=False):
                    grp.to_excel(writer, sheet_name=to_safe_sheet(agent), index=False)

                st.session_state["last_nom_df"].to_excel(writer, sheet_name="Nomination Support Details", index=False)
                st.session_state["last_result_df"].to_excel(writer, sheet_name="Comparison", index=False)
                st.session_state["last_nomination_df"].to_excel(writer, sheet_name="Nomination", index = False)

            buf.seek(0)
            st.download_button(
                "📥 Download Excel",
                data=buf.getvalue(),
                file_name="cif_charge_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # 7‑C Save Feature ----------------------------------------------------
        with save_placeholder:
            if "save_mode" not in st.session_state:
                st.session_state.save_mode = False

            if not st.session_state.save_mode:
                if st.button("💾 Save Comparison"):
                    st.session_state.save_mode = True
                    st.rerun()
            else:
                st.text_input("Enter a name for this comparison:", key="save_filename")
                confirm_col, cancel_col = st.columns([1, 1])
                if confirm_col.button("✅ Confirm Save"):
                    filename = st.session_state.get("save_filename", "").strip()
                    if not filename:
                        st.error("Filename cannot be empty.")
                    else:
                        safe_name = re.sub(r"[^A-Za-z0-9 _-]", "", filename).replace(" ", "_")
                        file_path = os.path.join(SAVED_DIR, f"{safe_name}.xlsx")
                        if os.path.exists(file_path):
                            st.warning("A file with this name already exists and will be overwritten.")
                        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
                            st.session_state["container_info"].to_excel(writer, sheet_name="Info", index=False)
                            for agent, grp in st.session_state["last_input_df"].groupby("Agent Name", sort=False):
                                grp.to_excel(writer, sheet_name=to_safe_sheet(agent), index=False)
                            st.session_state["last_nom_df"].to_excel(writer, sheet_name="Nomination Support Details", index=False)
                            st.session_state["last_result_df"].to_excel(writer, sheet_name="Comparison", index=False)
                            st.session_state["last_nomination_df"].to_excel(writer, sheet_name="Nomination", index = False)
                        st.success(f"Comparison saved as '{safe_name}.xlsx' in the Saved folder.")
                        st.session_state.save_mode = False
                if cancel_col.button("❌ Cancel"):
                    st.session_state.save_mode = False
                    st.rerun()
    else:
        with dl_placeholder:
            st.caption("Run **Calculate** first to enable download and save options.")

# ==============================================================================
# TAB 2: SAVED COMPARISONS
# ==============================================================================
import streamlit as st
import pandas as pd
import os
from io import BytesIO
import zipfile
import openpyxl

with main_tabs[1]:
    st.title("📂 Saved Comparisons")

    saved_files = [f for f in os.listdir(SAVED_DIR) if f.lower().endswith(".xlsx")]

    if not saved_files:
        st.info("No saved comparisons found. Return to the first tab, perform a calculation, and save it.")
    else:
        col1, col2 = st.columns([3, 1])
        with col1:
            selected_file = st.selectbox("Select a saved comparison to view:", saved_files)

        with col2:
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zipf:
                for f in saved_files:
                    zipf.write(os.path.join(SAVED_DIR, f), arcname=f)
            zip_buffer.seek(0)
            st.download_button(
                label="🛆 Download All",
                data=zip_buffer,
                file_name="All_Comparisons.zip",
                mime="application/zip"
            )

        if selected_file:
            file_path = os.path.join(SAVED_DIR, selected_file)

            if st.button("➕ Add New Agent Sheet"):
                wb = openpyxl.load_workbook(file_path)
                existing = wb.sheetnames
                base = "NewAgent"
                count = 1
                while f"{base}{count}" in existing:
                    count += 1
                new_sheet = f"{base}{count}"
                headers = ["Agent Name", "Description", "Currency", "Per CBM", "Per Ton", "Minimum", "Maximum", "Per BL", "Per Container"]
                ws = wb.create_sheet(title=new_sheet)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)
                wb.save(file_path)
                wb.close()
                st.success(f"✅ '{new_sheet}' added.")
                st.rerun()

            with open(file_path, "rb") as raw_file:
                file_bytes = BytesIO(raw_file.read())
            xl = pd.ExcelFile(file_bytes)
            sheet_names = xl.sheet_names

            nom_sheet_name = "Nomination Support Details"
            nom_df = xl.parse(nom_sheet_name) if nom_sheet_name in sheet_names else pd.DataFrame()

            special_sheets = ["Info", "Comparison", "Nomination", nom_sheet_name]
            agent_sheets = [s for s in sheet_names if s not in special_sheets]
            tab_order = ["Info"] if "Info" in sheet_names else []
            tab_order += agent_sheets
            if nom_sheet_name in sheet_names:
                tab_order.append(nom_sheet_name)
            if "Comparison" in sheet_names:
                tab_order.append("Comparison")
            if "Nomination" in sheet_names:
                tab_order.append("Nomination")

            view_tabs = st.tabs(tab_order)
            nom_dict = {}
            agents_data = {}

            for sheet, tab in zip(tab_order, view_tabs):
                with tab:
                    df = xl.parse(sheet)

                    if sheet in agent_sheets:
                        st.subheader(f"✏️ Edit Agent Sheet: {sheet}")

                        agent_name_val = df["Agent Name"].iloc[0] if "Agent Name" in df.columns and not df.empty else ""
                        agent_name_input = st.text_input("👤 Agent Name", value=agent_name_val, key=f"agent_name_{sheet}")

                        nom_row = nom_df[nom_df["Agent Name"] == agent_name_val] if not nom_df.empty else pd.DataFrame()
                        nom_rate = float(nom_row["Nomination Rate"].values[0]) if not nom_row.empty else 0.0
                        nom_cbm = float(nom_row["Nomination CBM"].values[0]) if not nom_row.empty else 0.0
                        nom_bl  = float(nom_row["Nomination BL"].values[0]) if not nom_row.empty else 0.0

                        st.markdown("### 📄 Nomination Support (Auto-linked)")
                        c1, c2, c3 = st.columns(3)
                        with c1:
                           nomination_rate = st.number_input("Nomination Rate", value=nom_rate, key=f"nom_rate_{sheet}")
                        with c2:
                            nomination_cbm = st.number_input("Nomination CBM", value=nom_cbm, key=f"nom_cbm_{sheet}")
                        with c3:
                            nomination_bl = st.number_input("Nomination BL", value=nom_bl, key=f"nom_bl_{sheet}")

                        nom_dict[agent_name_input] = {
                                "Agent Name": agent_name_input,
                                "Nomination Rate": nomination_rate,
                                "Nomination CBM": nomination_cbm,
                                "Nomination BL": nomination_bl
                            }

                        remarks_row = df[df["Description"] == "Remarks"].fillna("")
                        rebate_row = df[df["Description"] == "Rebate"].fillna(0)
                        editable_df = df[~df["Description"].isin(["Remarks", "Rebate"])].reset_index(drop=True)
                        if editable_df.empty:
                            default_row = {
                                "Description": "",
                                "Currency": currency_options[0],
                                "Per CBM": 0.0,
                                "Per Ton": 0.0,
                                "Minimum": 0.0,
                                "Maximum": 0.0,
                                "Per BL": 0.0
                            }
                            editable_df = pd.DataFrame([default_row])

                        columns_to_hide = ["Agent Name", "Per Container"]
                        editable_df_display = editable_df.drop(columns=[c for c in columns_to_hide if c in editable_df.columns])

                        st.markdown("### ✏️ Editable Charge Heads")

                        editable_df_display = editable_df_display.fillna("")

                        column_configs = {
                            "Currency": st.column_config.SelectboxColumn(
                                label="Currency",
                                options=currency_options,
                                required=True
                            )
                        }

                        agent_edited_df = st.data_editor(
                            editable_df_display,
                            use_container_width=True,
                            column_config=column_configs,
                            key=f"editor_{sheet}",
                            num_rows="dynamic"
                        )

                        remarks_text = remarks_row["Currency"].values[0] if not remarks_row.empty else ""
                        remarks_ = st.text_area("📒 Remarks", value=remarks_text or "", key=f"remarks_{sheet}")

                        rebate_currency = rebate_row["Currency"].values[0] if not rebate_row.empty else "USD"
                        rebate_cbm = rebate_row["Per CBM"].values[0] if not rebate_row.empty else 0.0
                        rebate_ton = rebate_row["Per Ton"].values[0] if not rebate_row.empty else 0.0
                        rebate_bl = rebate_row["Per BL"].values[0] if not rebate_row.empty else 0.0
                        rebate_container = rebate_row["Per Container"].values[0] if not rebate_row.empty else 0.0

                        st.markdown("### 💰 Rebate Details")
                        rc1, rc2, rc3, rc4, rc5 = st.columns(5)
                        with rc1:
                            st.selectbox(
                                "Currency",
                                options=currency_options,
                                index=currency_options.index(rebate_currency) if rebate_currency in currency_options else 0,
                                key=f"rebate_cur_{sheet}"
                            )
                        with rc2:
                            rebate_cbm = st.number_input("Per CBM", value=float(rebate_cbm), key=f"rebate_cbm_{sheet}")
                        with rc3:
                            rebate_ton = st.number_input("Per Ton", value=float(rebate_ton), key=f"rebate_ton_{sheet}")
                        with rc4:
                            rebate_bl = st.number_input("Per BL", value=float(rebate_bl), key=f"rebate_bl_{sheet}")
                        with rc5:
                            rebate_pcont = st.number_input("Per Container", value=float(rebate_container), key=f"rebate_con_{sheet}")

                        # Append remarks and rebate as new rows to the agent_edited_df
                        remarks_row_df = pd.DataFrame([{
                            "Agent Name": agent_name_input,
                            "Description": "Remarks",
                            "Currency": remarks_,
                            "Per CBM": "", "Per Ton": "", "Minimum": "", "Maximum": "", "Per BL": "", "Per Container": ""
                        }])

                        rebate_row_df = pd.DataFrame([{
                            "Description": "Rebate",
                            "Currency": st.session_state.get(f"rebate_cur_{sheet}", "USD"),
                            "Per CBM": rebate_cbm,
                            "Per Ton": rebate_ton,
                            "Minimum": "",  # optional, adjust if needed
                            "Maximum": "",
                            "Per BL": rebate_bl,
                            "Per Container": rebate_pcont
                        }])

                        # Add the two special rows to the edited DataFrame
                        agent_final_df = pd.concat([agent_edited_df, remarks_row_df, rebate_row_df], ignore_index=True)
                        agent_final_df["Agent Name"] = agent_name_input
                        # Store for later usage (e.g. for calculate, export, or save)
                        agents_data[agent_name_input] = agent_final_df


                    elif sheet == "Info":
                        st.subheader("✏️ Edit Info Sheet")

                        info_dict = df.set_index("Field").T.to_dict()

                        def get_val(field, col):
                            try:
                                return str(info_dict[field][col])
                            except:
                                return "0"

                        with st.expander("***📦 20' STD Information***", expanded=True):
                            c1, c2, c3, c4, c5 = st.columns(5)
                            box_rate_20    = c1.text_input("**Box Rate (USD)**", get_val("Box Rate (USD)", "20'STD"), key=f"box_rate_20_{sheet}")
                            loadability_20 = c2.text_input("**Loadability (numeric)**", get_val("Loadability", "20'STD"), key=f"load_20_{sheet}")
                            num_bl_20      = c3.text_input("**Number of BLs (numeric)**", get_val("Number of BLs", "20'STD"), key=f"num_bl_20_{sheet}")
                            market_rate_20 = c4.text_input("**Market Rate (USD)**", get_val("Market Rate (USD)", "20'STD"), key=f"mkt_rate_20_{sheet}")

                            try:
                                box_20 = float(box_rate_20)
                                load_20 = float(loadability_20)
                                if load_20 > 0:
                                    freight_cost_20 = box_20 / load_20
                                    c5.metric("📉 Freight Cost Per CBM", f"${freight_cost_20:.2f}")
                                else:
                                    c5.write("Enter loadability > 0")
                            except ValueError:
                                c5.write("Waiting for valid numbers")

                            st.markdown("**Transhipment**")
                            t1, t2, t3, _, _ = st.columns(5)
                            tran_cbm_20        = t1.text_input("**CBM (numeric)**", get_val("Transhipment CBM", "20'STD"), key=f"tran_cbm_20_{sheet}")
                            tran_num_bl_20     = t2.text_input("**# of BLs (numeric)**", get_val("Transhipment Number of BLs", "20'STD"), key=f"tran_num_bl_20_{sheet}")
                            tran_pro_per_cbm_20 = t3.text_input("**Profitability Per CBM**", get_val("Transhipment Profitability Per CBM", "20'STD"), key=f"tran_pro_per_cbm_20_{sheet}")

                        with st.expander("***📦 40' STD Information***", expanded=True):
                            c1, c2, c3, c4, c5 = st.columns(5)
                            box_rate_40    = c1.text_input("**Box Rate (USD)**", get_val("Box Rate (USD)", "40'STD"), key=f"box_rate_40_{sheet}")
                            loadability_40 = c2.text_input("**Loadability (numeric)**", get_val("Loadability", "40'STD"), key=f"load_40_{sheet}")
                            num_bl_40      = c3.text_input("**Number of BLs (numeric)**", get_val("Number of BLs", "40'STD"), key=f"num_bl_40_{sheet}")
                            market_rate_40 = c4.text_input("**Market Rate (USD)**", get_val("Market Rate (USD)", "40'STD"), key=f"mkt_rate_40_{sheet}")

                            try:
                                box_40 = float(box_rate_40)
                                load_40 = float(loadability_40)
                                if load_40 > 0:
                                    freight_cost_40 = box_40 / load_40
                                    c5.metric("📉 Freight Cost Per CBM", f"${freight_cost_40:.2f}")
                                else:
                                    c5.write("Enter loadability > 0")
                            except ValueError:
                                c5.write("Waiting for valid numbers")

                            st.markdown("**Transhipment**")
                            t1, t2, t3, _, _ = st.columns(5)
                            tran_cbm_40        = t1.text_input("**CBM (numeric)**", get_val("Transhipment CBM", "40'STD"), key=f"tran_cbm_40_{sheet}")
                            tran_num_bl_40     = t2.text_input("**# of BLs (numeric)**", get_val("Transhipment Number of BLs", "40'STD"), key=f"tran_num_bl_40_{sheet}")
                            tran_pro_per_cbm_40 = t3.text_input("**Profitability Per CBM**", get_val("Transhipment Profitability Per CBM", "40'STD"), key=f"tran_pro_per_cbm_40_{sheet}")
                    else:
                        st.subheader(f"🔍 Preview: {sheet}")
                        st.dataframe(df, use_container_width=True)

            col_re, col_dl, col_del = st.columns([1, 1, 1])
            if col_re.button("🧮 Re-Calculate"):
                try:
                    box_rate_20_f = float(box_rate_20)
                    loadability_20_f = float(loadability_20)
                    num_bl_20_f = float(num_bl_20)
                    market_rate_20_f = float(market_rate_20)
                    tran_cbm_20_f = float(tran_cbm_20)
                    tran_num_bl_20_f = float(tran_num_bl_20)
                    tran_pro_per_cbm_20_f = float(tran_pro_per_cbm_20)

                    box_rate_40_f = float(box_rate_40)
                    loadability_40_f = float(loadability_40)
                    num_bl_40_f = float(num_bl_40)
                    market_rate_40_f = float(market_rate_40)
                    tran_cbm_40_f = float(tran_cbm_40)
                    tran_num_bl_40_f = float(tran_num_bl_40)
                    tran_pro_per_cbm_40_f = float(tran_pro_per_cbm_40)

                except ValueError:
                    st.error("Loadability, Box Rate, and Origin Charges must be numeric.")
                    st.stop()

                input_dict = {
                        "20'STD": [
                            loadability_20_f, box_rate_20_f, num_bl_20_f, market_rate_20_f,
                            tran_cbm_20_f, tran_num_bl_20_f, tran_pro_per_cbm_20_f
                        ],
                        "40'STD": [
                            loadability_40_f, box_rate_40_f, num_bl_40_f, market_rate_40_f,
                            tran_cbm_40_f, tran_num_bl_40_f, tran_pro_per_cbm_40_f
                        ]
                }

                # Convert nom_dict to DataFrame
                nom_df = pd.DataFrame(nom_dict.values())

                # Combine all agent sheets into one DataFrame
                agent_dfs = []
                for agent_name, df in agents_data.items():
                    df = df.copy()
                    df["Agent Name"] = agent_name  # Ensure Agent Name column is consistent
                    agent_dfs.append(df)

                if agent_dfs:
                    in_df = pd.concat(agent_dfs, ignore_index=True)
                else:
                    in_df = pd.DataFrame()  # Empty fallback

                # Run the comparison
                comp_df, nomination_df = agent_compare(in_df, nom_df, input_dict, exchange_df)

                # Save to session
                st.session_state["container_info"] = pd.DataFrame({
                    "Field": [
                        "POL", "POD", "Loadability", "Box Rate (USD)", "Number of BLs", "Market Rate (USD)",
                        "Transhipment CBM", "Transhipment Number of BLs", "Transhipment Profitability Per CBM"
                    ],
                    "20'STD": [
                        pol, pod, loadability_20_f, box_rate_20_f, num_bl_20_f, market_rate_20_f,
                        tran_cbm_20_f, tran_num_bl_20_f, tran_pro_per_cbm_20_f
                    ],
                    "40'STD": [
                        pol, pod, loadability_40_f, box_rate_40_f, num_bl_40_f, market_rate_40_f,
                        tran_cbm_40_f, tran_num_bl_40_f, tran_pro_per_cbm_40_f
                    ]
                })

                st.session_state["last_input_df"] = in_df
                st.session_state["last_nom_df"] = nom_df
                st.session_state["last_result_df"] = comp_df
                st.session_state["last_nomination_df"] = nomination_df

                st.success("✅ Recalculation complete.")
                st.dataframe(comp_df)
                st.dataframe(nom_df)
                st.dataframe(nomination_df)

                # Overwrite the selected Excel file with all updated sheets
                with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
                    # Save container info
                    st.session_state["container_info"].to_excel(writer, sheet_name="Info", index=False)

                    # Save each agent sheet separately
                    for agent_name, df in agents_data.items():
                        safe_name = re.sub(r"[\[\]\*:/\\?]", "", agent_name)[:31] or "Sheet"
                        df.to_excel(writer, sheet_name=safe_name, index=False)

                    # Save updated Nomination Support Details
                    nom_df.to_excel(writer, sheet_name="Nomination Support Details", index=False)

                    # Save calculated results
                    st.session_state["last_result_df"].to_excel(writer, sheet_name="Comparison", index=False)
                    st.session_state["last_nomination_df"].to_excel(writer, sheet_name="Nomination", index=False)

                st.success(f"💾 Changes saved to '{selected_file}' successfully.")



            with col_dl:
                with open(file_path, "rb") as fp:
                    st.download_button(
                        label="📅 Download this comparison",
                        data=fp.read(),
                        file_name=selected_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            with col_del:
                if "delete_mode" not in st.session_state:
                    st.session_state.delete_mode = False
                if not st.session_state.delete_mode:
                    if st.button("🚨 Delete this comparison"):
                        st.session_state.delete_mode = True
                        st.rerun()
                else:
                    st.warning("Are you sure you want to delete this file? This action cannot be undone.")
                    c1, c2 = st.columns(2)
                    if c1.button("✅ Yes, delete"):
                        try:
                            os.remove(file_path)
                            st.success("Deleted successfully.")
                            st.session_state.delete_mode = False
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error deleting file: {e}")
                    if c2.button("❌ Cancel"):
                        st.session_state.delete_mode = False
                        st.rerun()



with main_tabs[2]:
    st.title("💱 Edit Exchange Rates")
    st.caption("You can update or add new exchange rates. Click save to apply changes.")

    edited_df = st.data_editor(
        exchange_df,
        num_rows="dynamic",
        use_container_width=True,
        key="exchange_editor"
    )

    if st.button("💾 Save Exchange Rates"):
        if "Currency" in edited_df.columns and "Exchange Rate to USD" in edited_df.columns:
            try:
                edited_df["Exchange Rate to USD"] = pd.to_numeric(edited_df["Exchange Rate to USD"])
                save_exchange_rates(edited_df)
                st.success("Exchange rates saved successfully. Please refresh to see changes.")
                st.cache_data.clear()
            except Exception as e:
                st.error(f"Error saving exchange rates: {e}")
        else:
            st.error("Please ensure 'Currency' and 'Exchange Rate to USD' columns exist.")

with main_tabs[3]:
    st.title("🚢 Edit Port Of Discharge")
    st.caption("You can update or add new PODs. Click save to apply changes.")

    # Load POD Excel Sheet
    pod_path = "Data/locations.xlsx"
    pod_df = pd.read_excel(pod_path, sheet_name="POD locations")

    # Data Editor
    edited_pod_df = st.data_editor(
        pod_df,
        num_rows="dynamic",
        use_container_width=True,
        key="pod_editor"
    )

    # Save Button
    if st.button("💾 Save PODs"):
        try:
            # Overwrite same sheet in the Excel file
            with pd.ExcelWriter(pod_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                edited_pod_df.to_excel(writer, sheet_name="POD locations", index=False)
            st.success("POD list saved successfully. Please refresh to see changes.")
        except Exception as e:
            st.error(f"Error saving PODs: {e}")
